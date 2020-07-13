
# coding: utf-8

# In[6]:


import os
import skimage.io

# Root directory of the project
ROOT_DIR = os.path.abspath("./")

frame = skimage.io.imread(os.path.join(ROOT_DIR, "dendaro3.jpg"))


# In[9]:


import openpyxl


# In[11]:


BookA = openpyxl.load_workbook(os.path.join(ROOT_DIR, "color_base.xlsx"))
SheetOUT = BookA.get_sheet_by_name('Red')
ndarrayA=frame[:,:,0].copy()
for rowA in range(ndarrayA.shape[0]):
    for colA in range(ndarrayA.shape[1]):
        SheetOUT.cell(row=rowA+1, column=colA+1, value=ndarrayA[rowA][colA])
BookA.save(os.path.join(ROOT_DIR, "color_out.xlsx"))


# In[ ]:


import openpyxl
BookA = openpyxl.load_workbook(os.path.join(ROOT_DIR, "color_base.xlsx"))
SheetOUT = BookA.get_sheet_by_name('Green')
ndarrayA=frame[:,:,1].copy()
for rowA in range(ndarrayA.shape[0]):
    for colA in range(ndarrayA.shape[1]):
        SheetOUT.cell(row=rowA+1, column=colA+1, value=ndarrayA[rowA][colA])
BookA.save(os.path.join(ROOT_DIR, "color_out.xlsx"))


# In[ ]:


import openpyxl
BookA = openpyxl.load_workbook(os.path.join(ROOT_DIR, "color_base.xlsx"))
SheetOUT = BookA.get_sheet_by_name('Blue')
ndarrayA=frame[:,:,2].copy()
for rowA in range(ndarrayA.shape[0]):
    for colA in range(ndarrayA.shape[1]):
        SheetOUT.cell(row=rowA+1, column=colA+1, value=ndarrayA[rowA][colA])
BookA.save(os.path.join(ROOT_DIR, "color_out.xlsx"))

